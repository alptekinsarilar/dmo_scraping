from bs4 import BeautifulSoup
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# set file path
filepath = "./dmo_urunler.xlsx"

wb = load_workbook(filepath)

# create new sheet
#wb.create_sheet('İş Makineleri')
#
#ws = wb['İş Makineleri']

# Specify the name of the worksheet you want to create or use
worksheet_name = 'Ekipmanlar'

# Check if the worksheet already exists in the workbook
if worksheet_name in wb.sheetnames:
    del wb[worksheet_name]
    ws = wb.create_sheet(worksheet_name)  # Create a new worksheet
    headings = ["Ürün Bilgisi", "Satış Koşulları", "Ürün Özellikleri"]
    ws.append(headings)
    wb.save(filepath)
    ws = wb[worksheet_name]  # Select the existing worksheet
else:
    ws = wb.create_sheet(worksheet_name)  # Create a new worksheet
    headings = ["Ürün Bilgisi", "Satış Koşulları", "Ürün Özellikleri"]
    ws.append(headings)
    wb.save(filepath)

def scrape_product(url):
    try:
        source = requests.get("https://www.dmo.gov.tr" + url)
        source.raise_for_status()
        soup = BeautifulSoup(source.text, "html.parser")

        #with open("ornek_urun.html", "r") as file:
        #    soup = BeautifulSoup(file, "html.parser")

        """ Part 1 """

        baslik = soup.find("div", class_="title").text
        tedarikci = soup.find("a", class_="available").text
        #aciklama = soup.find_all("p")[0].string
        #vergisiz_fiyat = soup.find("div", class_="price-current").text
        #vergili_fiyat = soup.find("div", class_="price-prev").text

        #birim_fiyat = soup.find("div", class_="excerpt").find_all("div")[8].text.strip()
        #kdv = soup.find("div", class_="excerpt").find_all("div")[9].text.strip()
        #dmo_urun_kodu = soup.find("div", class_="excerpt").find_all("div")[10].text.strip()
        #olcu_birimi = soup.find("div", class_="excerpt").find_all("div")[11].text.strip()
        #orijinal_urun_kodu = soup.find("div", class_="excerpt").find_all("div")[12].text.strip()
        #garanti_suresi = soup.find("div", class_="excerpt").find_all("div")[13].text.strip()
        #yerli_katki_orani = soup.find("div", class_="excerpt").find_all("div")[14].text.strip()
        #teknolojik_duzey = soup.find("div", class_="excerpt").find_all("div")[15].text.strip()
        #yerli_mali_belgesi_gecerlilik_tarihi = soup.find("div", class_="excerpt").find_all("div")[16].text.strip()

        #strings = [baslik, tedarikci, aciklama, vergisiz_fiyat, vergili_fiyat, birim_fiyat, kdv, dmo_urun_kodu, olcu_birimi, orijinal_urun_kodu, garanti_suresi, yerli_katki_orani, teknolojik_duzey, yerli_mali_belgesi_gecerlilik_tarihi]

        #genel_bilgi = ""
        #for s in strings:
        #    if s is not None:
        #        genel_bilgi += s + "\n"
        
        #print(baslik)
        #print(tedarikci)
        #print()
        #print(aciklama)
        #print(vergisiz_fiyat)
        #print(vergili_fiyat)

        #print(birim_fiyat)
        #print(kdv)
        #print(dmo_urun_kodu)
        #print(olcu_birimi)
        #print(orijinal_urun_kodu)
        #print(garanti_suresi)
        #print(yerli_katki_orani)
        #print(teknolojik_duzey)
        #print(yerli_mali_belgesi_gecerlilik_tarihi)
        
        genel_bilgi = baslik + tedarikci + soup.find("div", class_="excerpt").get_text()
        
        #print(genel_bilgi)

        """ Part 2 """

        satis_kosullari = soup.find("section", class_="fadeInUp").text

        #print(satis_kosullari)

        """ Part 3 """

        urun_ozellikleri = soup.find_all("section", class_="fadeInUp")[1].text

        #print(urun_ozellikleri)


        """ Excel Part """
        ws.append([genel_bilgi, satis_kosullari, urun_ozellikleri])
        
        wb.save(filepath)

    except Exception as e:
        print(e)
        